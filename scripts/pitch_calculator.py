"""Nadia + Marketing Pitch Calculator — Excel builder.

Creates a formatted .xlsx with inputs, scenarios, scaling table,
and a bar chart. Upload to Google Sheets for live editing.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).parent / "Nadia_Marketing_Pitch_Calculator.xlsx"

# Styles
HEADER_FILL = PatternFill(start_color="1A3366", end_color="1A3366", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
SUBHEADER_FILL = PatternFill(start_color="336699", end_color="336699", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill(start_color="FFF2B3", end_color="FFF2B3", fill_type="solid")
INPUT_FONT = Font(bold=True, size=11)
LABEL_FONT = Font(size=10)
VALUE_FONT = Font(size=10)
CALLOUT_FONT = Font(bold=True, size=11, color="CC3300")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SGD_FMT = '$#,##0'
PCT_FMT = '0%'
ROI_FMT = '0%'


def style_header_row(ws, row, max_col=6):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def style_subheader_row(ws, row, max_col=6):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def style_input(cell):
    cell.fill = INPUT_FILL
    cell.font = INPUT_FONT
    cell.border = THIN_BORDER


def style_label(cell):
    cell.font = LABEL_FONT
    cell.alignment = Alignment(horizontal="left")


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pitch Calculator"

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18

    # ══════════════════════════════════════════════════════════════
    # SECTION 1: INPUTS (rows 1-9)
    # ══════════════════════════════════════════════════════════════
    r = 1
    ws.cell(r, 1, "INPUTS").font = HEADER_FONT
    ws.cell(r, 3, "Change yellow cells live")
    style_header_row(ws, r)

    r = 2; ws.cell(r, 1, "Monthly Ad Spend (SGD)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, 1000); style_input(c); c.number_format = SGD_FMT

    r = 3; ws.cell(r, 1, "Cost per Lead (SGD)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, 240); style_input(c); c.number_format = SGD_FMT

    r = 4; ws.cell(r, 1, "Leads per Month"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2); c.value = "=B2/B3"; c.number_format = '0.0'

    r = 5; ws.cell(r, 1, "Lead to Appointment Rate (%)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, 0.50); style_input(c); c.number_format = PCT_FMT

    r = 6; ws.cell(r, 1, "Appointments from Leads"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2); c.value = "=B4*B5"; c.number_format = '0.0'

    r = 7; ws.cell(r, 1, "Cost per Appointment (SGD)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2); c.value = "=B2/B6"; c.number_format = SGD_FMT

    r = 8; ws.cell(r, 1, "Close Rate (%)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, 0.30); style_input(c); c.number_format = PCT_FMT

    r = 9; ws.cell(r, 1, "Avg Commission per Deal (SGD)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, 10000); style_input(c); c.number_format = SGD_FMT

    # ══════════════════════════════════════════════════════════════
    # SECTION 2: APPOINTMENT VALUE (rows 11-15)
    # ══════════════════════════════════════════════════════════════
    r = 11
    ws.cell(r, 1, "APPOINTMENT VALUE")
    style_header_row(ws, r)

    r = 12; ws.cell(r, 1, "Avg Commission per Deal (SGD)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2); c.value = "=B9"; c.number_format = SGD_FMT

    r = 13; ws.cell(r, 1, "Appointments per Deal (current)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, 5); style_input(c)

    r = 14; ws.cell(r, 1, "Value per Appointment (SGD)"); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2); c.value = "=B12/B13"; c.number_format = SGD_FMT
    c.font = Font(bold=True, size=12, color="006600")

    r = 15
    ws.cell(r, 1, "Each appt is worth this much, but you pay:").font = CALLOUT_FONT
    c = ws.cell(r, 2); c.value = "=B7"; c.number_format = SGD_FMT; c.font = CALLOUT_FONT

    # ══════════════════════════════════════════════════════════════
    # SECTION 3: CURRENT vs NADIA+MARKETING (rows 17-25)
    # ══════════════════════════════════════════════════════════════
    r = 17
    ws.cell(r, 1, "SCENARIO COMPARISON")
    ws.cell(r, 2, "Current")
    ws.cell(r, 3, "Nadia + Marketing")
    style_header_row(ws, r, 3)

    rows_s3 = [
        (18, "Monthly Ad Spend (SGD)", "=B2", "=B2", SGD_FMT, False),
        (19, "Cost per Appointment (SGD)", 600, 400, SGD_FMT, True),
        (20, "Appointments / Month", "=B18/B19", "=C18/C19", '0.0', False),
        (21, "Close Rate (%)", 0.20, 0.35, PCT_FMT, True),
        (22, "Deals / Month", "=B20*B21", "=C20*C21", '0.0', False),
        (23, "Revenue / Month (SGD)", "=B22*$B$9", "=C22*$B$9", SGD_FMT, False),
        (24, "Cost per Deal / CAC (SGD)", "=B18/B22", "=C18/C22", SGD_FMT, False),
        (25, "ROI (%)", "=(B23-B18)/B18", "=(C23-C18)/C18", ROI_FMT, False),
    ]

    for row, label, val_b, val_c, fmt, is_input in rows_s3:
        ws.cell(row, 1, label); style_label(ws.cell(row, 1))
        cb = ws.cell(row, 2, val_b); cb.number_format = fmt; cb.border = THIN_BORDER
        cc = ws.cell(row, 3, val_c); cc.number_format = fmt; cc.border = THIN_BORDER
        if is_input:
            style_input(cb)
            style_input(cc)

    # ══════════════════════════════════════════════════════════════
    # SECTION 4: SCALING TABLES (rows 27-37)
    # ══════════════════════════════════════════════════════════════

    # Current model
    r = 27
    ws.cell(r, 1, "SCALING: CURRENT MODEL")
    style_header_row(ws, r)

    r = 28
    for i, h in enumerate(["Ad Spend (SGD)", "Cost/Appt", "Appointments", "Close Rate", "Deals", "Revenue (SGD)"], 1):
        ws.cell(r, i, h)
    style_subheader_row(ws, r)

    current_scale = [
        (29, 1000),
        (30, 10000),
        (31, 20000),
    ]
    for row, spend in current_scale:
        ws.cell(row, 1, spend).number_format = SGD_FMT
        ws.cell(row, 2).value = "=B19"; ws.cell(row, 2).number_format = SGD_FMT
        ws.cell(row, 3).value = f"=A{row}/B{row}"; ws.cell(row, 3).number_format = '0.0'
        ws.cell(row, 4).value = "=B21"; ws.cell(row, 4).number_format = PCT_FMT
        ws.cell(row, 5).value = f"=C{row}*D{row}"; ws.cell(row, 5).number_format = '0.0'
        ws.cell(row, 6).value = f"=E{row}*$B$9"; ws.cell(row, 6).number_format = SGD_FMT
        for col in range(1, 7):
            ws.cell(row, col).border = THIN_BORDER

    # Nadia model
    r = 33
    ws.cell(r, 1, "SCALING: NADIA + MARKETING")
    style_header_row(ws, r)

    r = 34
    for i, h in enumerate(["Ad Spend (SGD)", "Cost/Appt", "Appointments", "Close Rate", "Deals", "Revenue (SGD)"], 1):
        ws.cell(r, i, h)
    style_subheader_row(ws, r)

    nadia_scale = [
        (35, 1000),
        (36, 10000),
        (37, 20000),
    ]
    for row, spend in nadia_scale:
        ws.cell(row, 1, spend).number_format = SGD_FMT
        ws.cell(row, 2).value = "=C19"; ws.cell(row, 2).number_format = SGD_FMT
        ws.cell(row, 3).value = f"=A{row}/B{row}"; ws.cell(row, 3).number_format = '0.0'
        ws.cell(row, 4).value = "=C21"; ws.cell(row, 4).number_format = PCT_FMT
        ws.cell(row, 5).value = f"=C{row}*D{row}"; ws.cell(row, 5).number_format = '0.0'
        ws.cell(row, 6).value = f"=E{row}*$B$9"; ws.cell(row, 6).number_format = SGD_FMT
        for col in range(1, 7):
            ws.cell(row, col).border = THIN_BORDER

    # ══════════════════════════════════════════════════════════════
    # SECTION 5: CHART DATA + BAR CHART (rows 39-42)
    # ══════════════════════════════════════════════════════════════
    r = 39
    ws.cell(r, 1, "CHART DATA")
    ws.cell(r, 2, "Current")
    ws.cell(r, 3, "Nadia + Marketing")
    style_header_row(ws, r, 3)

    ws.cell(40, 1, "Appointments / Month"); ws.cell(40, 2).value = "=B20"; ws.cell(40, 3).value = "=C20"
    ws.cell(41, 1, "Deals / Month"); ws.cell(41, 2).value = "=B22"; ws.cell(41, 3).value = "=C22"
    ws.cell(42, 1, "Revenue / Month (SGD)")
    ws.cell(42, 2).value = "=B23"; ws.cell(42, 2).number_format = SGD_FMT
    ws.cell(42, 3).value = "=C23"; ws.cell(42, 3).number_format = SGD_FMT

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Current vs Nadia + Marketing"
    chart.y_axis.title = "Count"
    chart.x_axis.title = ""
    chart.style = 10
    chart.width = 18
    chart.height = 12

    categories = Reference(ws, min_col=1, min_row=40, max_row=41)
    current_data = Reference(ws, min_col=2, min_row=39, max_row=41)
    nadia_data = Reference(ws, min_col=3, min_row=39, max_row=41)

    chart.add_data(current_data, titles_from_data=True)
    chart.add_data(nadia_data, titles_from_data=True)
    chart.set_categories(categories)

    # Color the series
    from openpyxl.chart.series import DataPoint
    chart.series[0].graphicalProperties.solidFill = "CC3333"  # red for current
    chart.series[1].graphicalProperties.solidFill = "339966"  # green for nadia

    ws.add_chart(chart, "A44")

    # Freeze top row for scrolling
    ws.freeze_panes = "A2"

    # Save
    wb.save(str(OUTPUT))
    print(f"\nFile saved: {OUTPUT}")
    print("Upload to Google Sheets: https://sheets.google.com → File → Import")


if __name__ == "__main__":
    build()
